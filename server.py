import cv2 as cv
import imagezmq
from imutils import build_montages
import face_recognition
import openpyxl as op
import numpy
import datetime
import imutils
import pickle
import argparse
import os
import socket
import zmq

TCP_IP = '192.168.86.23'    #IP address of the client
TCP_PORT = 5999     #portnumber
Message = 'DETECTED'



ap = argparse.ArgumentParser()
ap.add_argument("-e","--encodings",required=True,help="path to serialized db of facial encodings")
ap.add_argument("-o","--output",type=str,help="path to output video")
ap.add_argument("-y","--display",type=int,default=1,help="wheather or not to display output frame to screen")
ap.add_argument("-d","--detection-method",type=str,default="hog",help="face detection model to use: either hog or cnn")
ap.add_argument("-mW","--montageW",required=True,type=int,help="montage frame width")
ap.add_argument("-mH","--montageH",required=True,type=int,help="montage frame height")
args = vars(ap.parse_args()) 


#initialise the Imagehub object
imageHub = imagezmq.ImageHub()

# loading the known faces and embeddings
print("[INFO] loading encodings...")
data = pickle.loads(open(args["encodings"], "rb").read())

#Information regarding when a device was last active
lastActive = {}
lastActiveCheck = datetime.datetime.now()


mW = args["montageW"]
mH = args["montageH"]

name = "Unknown"
names = []
face_locations = []
face_encodings = []
totalframes = 1
totaldays = 0
m = 2
flag = 0


now= datetime.datetime.now()
today=now.day
month=now.month

face_cascade=cv.CascadeClassifier("haarcascade_frontalface_default.xml")
while True:
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    s.connect((TCP_IP, TCP_PORT))   #Connecting to the client
    totalframes = totalframes + 1
    (rpiName , frame) = imageHub.recv_image()   #Receiving image from the client
    imageHub.send_reply(b'OK')
    #cv.imshow("Video",frame)
    if rpiName not in lastActive.keys():
        print("[INFO] receiving data from {}....".format(rpiName))

    lastActive[rpiName] = datetime.datetime.now()

    rgb_frame = frame[: , : , ::-1]     #Converting the frame from BGR to RGB

    face_locations = face_recognition.face_locations(rgb_frame , model=args["detection_method"])
    face_encodings = face_recognition.face_encodings(rgb_frame,face_locations)    #encoding the grabbed frame


    gray = cv.cvtColor(frame,cv.COLOR_BGR2GRAY)     #Converting the frame from BGR to Gray
    faces = face_cascade.detectMultiScale(gray, 1.05 , 5)    #applying cascade classifier

    for x,y,w,h in faces:
        gray = cv.rectangle(gray,(x,y),(x+w,y+h),(0,255,0),3)

    cv.imshow("Video",gray)   #Feed display
    for face_encoding in face_encodings:
        data = pickle.loads(open(args["encodings"], "rb").read())     #Loading the already saved encodings
        matches = face_recognition.compare_faces(data["encodings"],face_encoding,0.4) #Comparing the current frame encoding with the already saved encoding
        if True in matches:
            matchedIds = [i for (i,b) in enumerate(matches) if b]
            for i in matchedIds:
                name = data["names"][i]     #grabbing the name of the encoding file with which the face is matched
                print(name)
                s.send(Message.encode())   #Sending the message to the client that the name is detected
                data = s.recv(20).decode()  #Reply of the sent request
                s.close()

#Storing the necessary information in the excel sheet

                book = op.load_workbook("INFORMATION.xlsx")
                book1 = op.load_workbook("attendance.xlsx")
                book2 = op.load_workbook("presentdetails.xlsx")


                sheet = book.active    #INFORMATION SHEET
                sheet1 = book1.active  #IN-TIME , OUT-TIME SHEET
                sheet2 = book2.active  #PRESENT-DAYS COUNT SHEET

                for j in range(2,40):
                    if sheet.cell( row = j ,column = 2 ).value == name:
                        sheet1.cell(row = m , column = 2).value = name
                        sheet1.cell(row = m , column = 1).value = sheet.cell(row = j , column = 1).value
                        sheet1.cell(row = m , column = 3).value = sheet.cell(row = j , column = 3).value
                        sheet1.cell(row = m , column = 4).value = datetime.datetime.now()
                        m = m + 1

                        sheet2.cell( row = j , column = int(today) + 4 ).value = "PRESENT"
                        for k in range (4,40):
                            if sheet2.cell(row = j , column = k).value == "PRESENT":
                                totaldays = totaldays + 1
                        sheet2.cell(row = j , column = 3).value = totaldays 
                        break

                book1.save("attendance.xlsx")   #Saving the files
                book2.save("presentdetails.xlsx")


    totaldays = 0
    name = "Unknown"
    flag = 0

    key = cv.waitKey(1)
    if  key == ord('q'):
        break

